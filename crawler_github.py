#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
校园招聘与实习信息爬虫工具

功能：
- 爬取指定网站的校招和实习岗位信息
- 筛选2026届相关职位
- 保存数据到JSON和Excel（新增职位高亮）
- 自动发送邮件通知新职位信息

使用前请配置环境变量：
- EMAIL_USER: 发送邮箱账号
- EMAIL_PWD: 发送邮箱授权码
- RECEIVER_EMAILS: 接收邮箱列表（分号分隔）
"""

import os
import time
import json
import logging
import smtplib
import random
import pandas as pd
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from fake_useragent import UserAgent
import openpyxl
import html

# ============================
# 配置与初始化
# ============================

# 配置日志系统
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# 常量定义
START_PAGE = 1  # 起始页码
END_PAGE = 6    # 目标总页码
MAX_PAGES_PER_SESSION = 2  # 每次会话最大爬取页数（反反爬策略）
WAIT_TIME_MIN = 1  # 页面等待最小时间（秒）
WAIT_TIME_MAX = 3  # 页面等待最大时间（秒）

# 目标网站URL
SITE_URL = "https://www.givemeoc.com"  # 校招岗位页面
SITE_URL_INTERNSHIP = "https://www.givemeoc.com/internship"  # 实习岗位页面

# 数据存储路径
DATA_FILE_CAMPUS = "campus_jobs.json"       # 校招数据JSON文件
DATA_FILE_INTERNSHIP = "intern_jobs.json"   # 实习数据JSON文件
EXCEL_FILE_CAMPUS = "campus_jobs.xlsx"      # 校招数据Excel文件
EXCEL_FILE_INTERNSHIP = "intern_jobs.xlsx"  # 实习数据Excel文件

# 邮箱配置（从环境变量获取）
EMAIL_USER = os.environ.get('EMAIL_USER')  # 发送邮箱账号
EMAIL_PWD = os.environ.get('EMAIL_PWD')    # 发送邮箱授权码
RECEIVER_EMAILS = os.environ.get('RECEIVER_EMAILS', '').split(';')  # 接收邮箱列表


# ============================
# 工具函数
# ============================

def is_target_recruitment(target):
    """判断是否为2026届相关招聘
    
    参数:
        target: 招聘对象描述文本
        
    返回:
        bool: 若包含2026届相关关键词则返回True，否则False
    """
    if not target:
        return False
    target_lower = target.lower()
    return '2026' in target_lower or '26届' in target_lower


def safe_get_text(element, selector):
    """安全获取元素文本（避免因元素不存在导致报错）
    
    参数:
        element: 父元素对象
        selector: CSS选择器
        
    返回:
        str: 元素文本（若获取失败则返回空字符串）
    """
    try:
        return element.find_element("css selector", selector).text
    except:
        return ""


def safe_get_attr(element, selector, attribute):
    """安全获取元素属性（避免因元素不存在导致报错）
    
    参数:
        element: 父元素对象
        selector: CSS选择器
        attribute: 要获取的属性名
        
    返回:
        str: 元素属性值（若获取失败则返回空字符串）
    """
    try:
        return element.find_element("css selector", selector).get_attribute(attribute)
    except:
        return ""


# ============================
# 数据处理函数
# ============================

def load_and_clean_historical_data(data_file):
    """加载并清理历史数据（仅保留2026届相关职位）
    
    参数:
        data_file: 历史数据JSON文件路径
        
    返回:
        dict: 清理后的历史数据字典，结构为:
            {
                "last_update": 最后更新时间ISO字符串,
                "jobs": {职位ID: 职位详情字典}
            }
    """
    try:
        if os.path.exists(data_file):
            with open(data_file, 'r', encoding='utf-8') as f:
                historical_data = json.load(f)
            
            # 筛选保留2026届相关职位
            original_count = len(historical_data.get("jobs", {}))
            filtered_jobs = {}
            for job_id, job in historical_data.get("jobs", {}).items():
                if is_target_recruitment(job.get("target", "")):
                    filtered_jobs[job_id] = job
            
            # 更新历史数据
            historical_data["jobs"] = filtered_jobs
            historical_data["last_clean_time"] = datetime.now().isoformat()
            
            # 记录清理结果
            removed_count = original_count - len(filtered_jobs)
            if removed_count > 0:
                logger.info(f"清理 {data_file}: 移除 {removed_count} 条非2026届职位，保留 {len(filtered_jobs)} 条")
                with open(data_file, 'w', encoding='utf-8') as f:
                    json.dump(historical_data, f, ensure_ascii=False, indent=2)
            else:
                logger.info(f"{data_file} 所有 {original_count} 条均为2026届相关职位")
            
            return historical_data
        else:
            logger.info(f"首次运行，初始化数据文件: {data_file}")
            return {"last_update": None, "jobs": {}}
    except Exception as e:
        logger.warning(f"加载历史数据失败，创建新数据集: {e}")
        return {"last_update": None, "jobs": {}}


def save_historical_data(data, data_file):
    """保存数据到本地JSON文件
    
    参数:
        data: 要保存的数据字典
        data_file: 目标文件路径
        
    返回:
        bool: 保存成功返回True，否则False
    """
    try:
        with open(data_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        logger.info(f"数据已保存至: {data_file}")
        return True
    except Exception as e:
        logger.error(f"保存数据失败: {e}")
        return False


def clean_expired_jobs(historical_data):
    """清理过期职位（截止日期已过）和非2026届职位
    
    参数:
        historical_data: 原始历史数据字典
        
    返回:
        dict: 清理后的历史数据字典
    """
    logger.info("开始清理过期职位...")
    current_time = datetime.now()
    expired_count = 0
    non_target_count = 0
    jobs_to_keep = {}
    
    for job_id, job in historical_data['jobs'].items():
        # 过滤非2026届职位
        if not is_target_recruitment(job.get("target", "")):
            non_target_count += 1
            continue
            
        # 过滤过期职位
        if job.get('deadline'):
            try:
                deadline_date = datetime.strptime(job['deadline'], "%Y-%m-%d")
                if deadline_date < current_time:
                    expired_count += 1
                    continue
            except:
                pass  # 无法解析的日期视为未过期
                
        jobs_to_keep[job_id] = job
    
    historical_data['jobs'] = jobs_to_keep
    logger.info(f"清理完成: 移除 {expired_count} 条过期职位，{non_target_count} 条非2026届职位，保留 {len(jobs_to_keep)} 条有效职位")
    return historical_data


# ============================
# 爬虫核心函数
# ============================

def setup_browser():
    """配置浏览器实例（带反反爬策略）
    
    返回:
        webdriver.Chrome: 配置好的浏览器驱动实例
    """
    chrome_options = Options()
    # 基础配置
    chrome_options.add_argument('--headless')  # 无头模式（无界面运行）
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument("--incognito")  # 无痕模式
    
    # 反反爬配置
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")  # 隐藏自动化特征
    chrome_options.binary_location = "/usr/bin/chromium-browser"  # GitHub Actions兼容
    
    # 随机User-Agent
    ua = UserAgent()
    chrome_options.add_argument(f"user-agent={ua.random}")
    
    # 初始化驱动
    driver = webdriver.Chrome(options=chrome_options)
    
    # 进一步隐藏自动化特征
    driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
        'source': 'Object.defineProperty(navigator, "webdriver", {get: () => undefined})'
    })
    
    logger.info("浏览器实例初始化完成")
    return driver


def crawl_campus_data(driver, site_url, start_page, end_page):
    """爬取校招页面数据（筛选2026届相关职位）
    
    参数:
        driver: 浏览器驱动实例
        site_url: 校招页面基础URL
        start_page: 起始页码
        end_page: 目标结束页码
        
    返回:
        tuple: (爬取的职位列表, 实际爬取的最后页码)
    """
    try:
        driver.get(site_url)
        time.sleep(random.uniform(WAIT_TIME_MIN, WAIT_TIME_MAX))
        
        # 跳转到起始页（若不是第1页）
        if start_page > 1:
            try:
                logger.info(f"跳转到校招第 {start_page} 页")
                page_input = driver.find_element("css selector", "input.crt-page-input")
                page_input.clear()
                page_input.send_keys(str(start_page))
                go_button = driver.find_element("css selector", "button.crt-page-go-btn")
                driver.execute_script("arguments[0].click();", go_button)  # 避免被检测为自动化点击
                time.sleep(random.gauss(3, 1))  # 高斯分布等待（更接近人类行为）
            except Exception as e:
                logger.error(f"校招跳转至第 {start_page} 页失败: {e}")
                return [], start_page - 1

        crawled_data = []
        current_page = start_page

        # 爬取当前会话分配的页数（最多2页）
        for page in range(start_page, min(end_page + 1, start_page + MAX_PAGES_PER_SESSION)):
            logger.info(f"爬取校招第 {page} 页")
            current_page = page

            # 模拟人类滚动
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(random.uniform(1, 2))

            # 解析职位列表
            job_items = driver.find_elements("css selector", "table.crt-table tbody tr")
            for item in job_items:
                try:
                    # 提取职位信息
                    job_info = {
                        "job_type": "校招",
                        "company": safe_get_text(item, "td.crt-col-company"),
                        "company_type": safe_get_text(item, "td.crt-col-type"),
                        "location": safe_get_text(item, "td.crt-col-recruitment-type + td"),
                        "recruitment_type": safe_get_text(item, "td.crt-col-recruitment-type"),
                        "target": safe_get_text(item, "td.crt-col-target"),
                        "position": safe_get_text(item, "td.crt-col-position"),
                        "update_time": safe_get_text(item, "td.crt-col-update-time"),
                        "deadline": safe_get_text(item, "td.crt-col-deadline"),
                        "links": safe_get_attr(item, "td.crt-col-links a", "href"),
                        "notice": safe_get_attr(item, "td.crt-col-notice a", "href"),
                        "referral": safe_get_text(item, "td.crt-col-referral"),
                        "notes": safe_get_text(item, "td.crt-col-notes"),
                        "crawl_time": datetime.now().isoformat()
                    }

                    # 筛选2026届相关职位
                    if is_target_recruitment(job_info["target"]):
                        crawled_data.append(job_info)
                    else:
                        logger.debug(f"过滤非2026届职位: {job_info['company']} - {job_info['position']}")
                except Exception as e:
                    logger.warning(f"处理校招职位失败: {e}")
                    continue

            # 翻到下一页（若未达当前会话终点）
            if page < min(end_page, start_page + MAX_PAGES_PER_SESSION - 1):
                try:
                    page_input = driver.find_element("css selector", "input.crt-page-input")
                    page_input.clear()
                    page_input.send_keys(str(page + 1))
                    go_button = driver.find_element("css selector", "button.crt-page-go-btn")
                    driver.execute_script("arguments[0].click();", go_button)
                    time.sleep(random.gauss(3, 1))
                except Exception as e:
                    logger.warning(f"校招翻页失败: {e}")
                    break

        return crawled_data, current_page
    except Exception as e:
        logger.error(f"校招爬取失败: {e}")
        return [], start_page


def crawl_internship_data(driver, site_url, start_page, end_page):
    """爬取实习页面数据（筛选2026届相关职位）
    
    参数:
        driver: 浏览器驱动实例
        site_url: 实习页面基础URL
        start_page: 起始页码
        end_page: 目标结束页码
        
    返回:
        tuple: (爬取的职位列表, 实际爬取的最后页码)
    """
    try:
        driver.get(site_url)
        time.sleep(random.uniform(WAIT_TIME_MIN, WAIT_TIME_MAX))
        
        # 跳转到起始页（若不是第1页）
        if start_page > 1:
            try:
                logger.info(f"跳转到实习第 {start_page} 页")
                page_input = driver.find_element("css selector", "input.int-page-input")
                page_input.clear()
                page_input.send_keys(str(start_page))
                go_button = driver.find_element("css selector", "button.int-page-go-btn")
                driver.execute_script("arguments[0].click();", go_button)
                time.sleep(random.gauss(3, 1))
            except Exception as e:
                logger.error(f"实习跳转至第 {start_page} 页失败: {e}")
                return [], start_page - 1

        crawled_data = []
        current_page = start_page

        # 爬取当前会话分配的页数（最多2页）
        for page in range(start_page, min(end_page + 1, start_page + MAX_PAGES_PER_SESSION)):
            logger.info(f"爬取实习第 {page} 页")
            current_page = page

            # 模拟人类滚动
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(random.uniform(1, 2))

            # 解析职位列表
            job_items = driver.find_elements("css selector", "table.int-table tbody tr")
            for item in job_items:
                try:
                    # 提取职位信息
                    job_info = {
                        "job_type": "实习",
                        "company": safe_get_text(item, "td.int-col-company"),
                        "company_type": safe_get_text(item, "td.int-col-type"),
                        "location": safe_get_text(item, "td.int-col-recruitment-type + td"),
                        "recruitment_type": safe_get_text(item, "td.int-col-recruitment-type"),
                        "target": safe_get_text(item, "td.int-col-target"),
                        "position": safe_get_text(item, "td.int-col-position"),
                        "update_time": safe_get_text(item, "td.int-col-update-time"),
                        "deadline": safe_get_text(item, "td.int-col-deadline"),
                        "links": safe_get_attr(item, "td.int-col-links a", "href"),
                        "notice": safe_get_attr(item, "td.int-col-notice a", "href"),
                        "referral": safe_get_text(item, "td.int-col-referral"),
                        "notes": safe_get_text(item, "td.int-col-notes"),
                        "crawl_time": datetime.now().isoformat()
                    }

                    # 筛选2026届相关职位
                    if is_target_recruitment(job_info["target"]):
                        crawled_data.append(job_info)
                    else:
                        logger.debug(f"过滤非2026届职位: {job_info['company']} - {job_info['position']}")
                except Exception as e:
                    logger.warning(f"处理实习职位失败: {e}")
                    continue

            # 翻到下一页（若未达当前会话终点）
            if page < min(end_page, start_page + MAX_PAGES_PER_SESSION - 1):
                try:
                    page_input = driver.find_element("css selector", "input.int-page-input")
                    page_input.clear()
                    page_input.send_keys(str(page + 1))
                    go_button = driver.find_element("css selector", "button.int-page-go-btn")
                    driver.execute_script("arguments[0].click();", go_button)
                    time.sleep(random.gauss(3, 1))
                except Exception as e:
                    logger.warning(f"实习翻页失败: {e}")
                    break

        return crawled_data, current_page
    except Exception as e:
        logger.error(f"实习爬取失败: {e}")
        return [], start_page


# ============================
# 输出与通知函数
# ============================

def save_excel_file(job_list, filename, added_jobs=None):
    """将职位数据保存为Excel（新增职位高亮）
    
    参数:
        job_list: 职位列表
        filename: 目标Excel路径
        added_jobs: 新增职位列表（用于高亮标记）
        
    返回:
        bool: 保存成功返回True，否则False
    """
    try:
        # 列名映射（中文显示）
        CN_HEADERS = {
            "company": "公司名称",
            "company_type": "公司类型",
            "location": "工作地点",
            "recruitment_type": "招聘类型",
            "target": "招聘对象",
            "position": "职位名称",
            "update_time": "更新时间",
            "deadline": "截止时间",
            "links": "职位链接",
            "notice": "通知链接",
            "referral": "内推信息",
            "notes": "备注",
            "crawl_time": "爬取时间"
        }
        
        # 筛选2026届相关职位
        filtered_jobs = [job for job in job_list if is_target_recruitment(job.get("target", ""))]
        df = pd.DataFrame(filtered_jobs).rename(columns=CN_HEADERS)
        
        # 标记新增职位
        if added_jobs:
            valid_added_ids = {f"{j['company']}-{j['position']}" for j in added_jobs 
                              if is_target_recruitment(j.get("target", ""))}
            df['_is_new'] = df.apply(
                lambda x: "是" if f"{x['公司名称']}-{x['职位名称']}" in valid_added_ids else "否", 
                axis=1
            )
        
        # 写入Excel并高亮新增职位
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='招聘信息')
            worksheet = writer.sheets['招聘信息']
            
            # 高亮处理
            if added_jobs and len(valid_added_ids) > 0:
                yellow_fill = openpyxl.styles.PatternFill(start_color="FFFF00", fill_type="solid")
                for row in worksheet.iter_rows(min_row=2):
                    if row[-1].value == "是":
                        for cell in row[:-1]:
                            cell.fill = yellow_fill
                worksheet.delete_cols(worksheet.max_column)  # 删除标记列
            
            # 调整列宽
            for col in worksheet.columns:
                max_len = max(len(str(cell.value)) for cell in col)
                worksheet.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)
        
        logger.info(f"Excel已保存至 {filename}，包含 {len(filtered_jobs)} 条2026届相关职位")
        return True
    except Exception as e:
        logger.error(f"保存Excel失败: {e}")
        return False


def generate_email_html(new_jobs, job_type):
    """生成美化的HTML邮件内容
    
    参数:
        new_jobs: 新增职位列表
        job_type: 职位类型（校招/实习）
        
    返回:
        str: HTML格式的邮件内容
    """
    # 筛选2026届相关职位
    filtered_jobs = [job for job in new_jobs if is_target_recruitment(job.get("target", ""))]
    
    # CSS样式（美化邮件）
    styles = """
    <style>
        body { font-family: 'Segoe UI', sans-serif; line-height: 1.6; color: #333; max-width: 800px; margin: 0 auto; padding: 20px; background: #f5f7fa; }
        .header { background: linear-gradient(135deg, #4b6cb7 0%, #182848 100%); color: white; padding: 20px; border-radius: 8px 8px 0 0; text-align: center; margin-bottom: 25px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }
        .header h1 { margin: 0; font-weight: 600; font-size: 24px; }
        .notification-card { background: white; border-radius: 8px; padding: 30px; margin-bottom: 30px; box-shadow: 0 4px 15px rgba(0,0,0,0.08); border: 1px solid #eaeaea; }
        .stats { display: flex; justify-content: space-around; margin-bottom: 25px; text-align: center; }
        .stat-item { background: #f0f5ff; padding: 15px; border-radius: 8px; flex: 1; margin: 0 10px; box-shadow: 0 2px 4px rgba(0,0,0,0.05); }
        .stat-item span { display: block; font-size: 28px; font-weight: bold; color: #4b6cb7; margin-bottom: 5px; }
        .job-item { background: #fff; border-left: 4px solid #4b6cb7; margin-bottom: 15px; padding: 15px; border-radius: 0 6px 6px 0; transition: all 0.3s ease; }
        .job-item:hover { transform: translateY(-3px); box-shadow: 0 5px 15px rgba(75, 108, 183, 0.15); }
        .company { font-weight: bold; color: #2c3e50; font-size: 18px; margin-bottom: 5px; }
        .position { font-weight: 600; color: #4b6cb7; font-size: 16px; margin: 10px 0; }
        .meta { display: flex; flex-wrap: wrap; gap: 15px; margin: 10px 0; color: #555; font-size: 14px; }
        .meta span:before { content: "•"; margin-right: 5px; color: #4b6cb7; }
        .deadline { background: #fff9e6; color: #e67e22; padding: 5px 10px; border-radius: 4px; font-weight: 600; display: inline-block; margin-top: 5px; }
        .target-highlight { background: #e3f2fd; color: #0d47a1; padding: 2px 5px; border-radius: 3px; font-weight: 500; }
        .links a { display: inline-block; background: #4b6cb7; color: white; text-decoration: none; padding: 8px 15px; border-radius: 4px; margin-top: 10px; transition: background 0.3s; }
        .links a:hover { background: #3a559f; }
        .notes { margin-top: 10px; padding: 10px; background: #f8f9fa; border-left: 3px solid #4b6cb7; font-size: 14px; color: #555; }
        .footer { text-align: center; margin-top: 30px; color: #777; font-size: 13px; padding: 15px; border-top: 1px solid #eee; }
    </style>
    """
    
    # 构建HTML内容
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>新职位通知 - {job_type}</title>
        {styles}
    </head>
    <body>
        <div class="header">
            <h1>🎯 新职位通知 - {job_type} (2026届相关)</h1>
        </div>
        
        <div class="notification-card">
            <div class="stats">
                <div class="stat-item">
                    <span>{len(filtered_jobs)}</span>
                    新职位
                </div>
                <div class="stat-item">
                    <span>{len(set(job['company'] for job in filtered_jobs))}</span>
                    家公司
                </div>
                <div class="stat-item">
                    <span>{datetime.now().strftime('%m/%d')}</span>
                    更新日期
                </div>
            </div>
            
            <div class="job-list">
    """
    
    # 添加职位列表
    for job in filtered_jobs:
        deadline = job.get('deadline', '截止时间待定')
        target = job.get('target', '')
        links_html = f'<div class="links"><a href="{job["links"]}" target="_blank">查看职位详情</a></div>' if job.get('links') else ""
        notes = f'<div class="notes">💡 职位亮点: {html.escape(job.get("notes", ""))}</div>' if job.get('notes') else ""
        
        html_content += f"""
        <div class="job-item">
            <div class="company">{html.escape(job.get('company', ''))}</div>
            <div class="position">🏢 {html.escape(job.get('position', ''))}</div>
            <div class="meta">
                <span>📍 {html.escape(job.get('location', ''))}</span>
                <span>🚀 {html.escape(job.get('recruitment_type', ''))}</span>
                <span>🎯 <span class="target-highlight">{html.escape(target)}</span></span>
            </div>
            <div class="deadline">⏰ 截止时间: {html.escape(str(deadline))}</div>
            {notes}
            {links_html}
        </div>
        """
    
    # 无新职位提示
    if not filtered_jobs:
        html_content += """
        <div class="no-jobs">
            <p>本次未发现符合条件的新职位（仅限2026届相关）。</p>
        </div>
        """
    
    # 邮件底部
    html_content += f"""
            </div>
        </div>
        <div class="footer">
            <p>自动爬虫系统生成 | 抓取时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            <p>© {datetime.now().year} 职位监控系统 | 共发现 {len(filtered_jobs)} 个2026届相关新职位</p>
        </div>
    </body>
    </html>
    """
    return html_content


def send_email(subject, body, attachment_paths=None):
    """发送邮件通知
    
    参数:
        subject: 邮件主题
        body: 邮件内容（HTML格式）
        attachment_paths: 附件路径列表（可选）
        
    返回:
        bool: 发送成功返回True，否则False
    """
    try:
        # 检查邮箱配置
        if not EMAIL_USER or not EMAIL_PWD or not RECEIVER_EMAILS:
            logger.warning("邮箱配置不完整，跳过发送")
            return False
            
        # 邮件服务器配置（QQ邮箱示例）
        smtp_server = "smtp.qq.com"
        smtp_port = 587

        # 构建邮件
        msg = MIMEMultipart()
        msg['From'] = EMAIL_USER
        msg['To'] = ", ".join(RECEIVER_EMAILS)
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'html'))

        # 添加附件
        if attachment_paths:
            for path in attachment_paths:
                if os.path.exists(path):
                    with open(path, 'rb') as f:
                        part = MIMEApplication(f.read(), Name=os.path.basename(path))
                    part['Content-Disposition'] = f'attachment; filename="{os.path.basename(path)}"'
                    msg.attach(part)

        # 发送邮件
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(EMAIL_USER, EMAIL_PWD)
            server.sendmail(EMAIL_USER, RECEIVER_EMAILS, msg.as_string())
        
        logger.info(f"邮件已发送至: {', '.join(RECEIVER_EMAILS)}")
        return True
    except Exception as e:
        logger.error(f"邮件发送失败: {e}")
        return False


# ============================
# 流程控制函数
# ============================

def process_site(site_name, site_url, data_file, excel_file):
    """处理单个站点的完整爬取流程
    
    参数:
        site_name: 站点名称（校招/实习）
        site_url: 站点URL
        data_file: 数据存储JSON路径
        excel_file: Excel输出路径
        
    返回:
        list: 所有有效职位列表（2026届相关）
    """
    logger.info(f"开始处理 {site_name} 站点（{START_PAGE}-{END_PAGE}页，每次{MAX_PAGES_PER_SESSION}页）")
    
    # 加载历史数据
    historical_data = load_and_clean_historical_data(data_file)
    existing_jobs = historical_data.get("jobs", {})
    all_new_jobs = []  # 累计所有新职位
    
    # 分页爬取（每次会话爬取2页，避免被反爬）
    current_start_page = START_PAGE
    while current_start_page <= END_PAGE:
        # 每次会话重新初始化浏览器
        driver = setup_browser()
        
        # 计算当前会话爬取范围
        current_end_page = min(current_start_page + MAX_PAGES_PER_SESSION - 1, END_PAGE)
        logger.info(f"=== 爬取 {current_start_page}-{current_end_page} 页 ===")
        
        # 调用对应爬虫函数
        if site_name == "校招":
            new_jobs, last_page = crawl_campus_data(driver, site_url, current_start_page, current_end_page)
        else:
            new_jobs, last_page = crawl_internship_data(driver, site_url, current_start_page, current_end_page)
        
        # 关闭浏览器
        driver.quit()
        
        # 处理新职位
        for job in new_jobs:
            job_id = f"{job['company']}-{job['position']}"  # 生成唯一ID
            if job_id not in existing_jobs:
                all_new_jobs.append(job)
                existing_jobs[job_id] = job
                logger.info(f"发现新职位: {job['company']} - {job['position']}")
        
        # 更新下一轮起始页
        current_start_page = last_page + 1
        logger.info(f"=== 完成 {current_start_page - MAX_PAGES_PER_SESSION}-{last_page} 页爬取 ===")
        
        # 爬取间隔（反反爬）
        if current_start_page <= END_PAGE:
            sleep_time = random.uniform(5, 10)
            logger.info(f"等待 {sleep_time:.1f} 秒后继续爬取...")
            time.sleep(sleep_time)
    
    # 更新并保存历史数据
    historical_data["jobs"] = existing_jobs
    historical_data["last_update"] = datetime.now().isoformat()
    historical_data = clean_expired_jobs(historical_data)  # 清理过期职位
    save_historical_data(historical_data, data_file)
    
    # 生成Excel并发送通知
    logger.info(f"{site_name} 爬取完成，新增 {len(all_new_jobs)} 个2026届相关职位")
    if save_excel_file(list(existing_jobs.values()), excel_file, added_jobs=all_new_jobs):
        email_body = generate_email_html(all_new_jobs, site_name)
        send_email(
            subject=f"{site_name}招聘信息更新（2026届相关）- {datetime.now().strftime('%Y%m%d')}",
            body=email_body,
            attachment_paths=[excel_file]
        )
    else:
        send_email(
            subject=f"{site_name}招聘信息更新（2026届相关）- {datetime.now().strftime('%Y%m%d')}",
            body=f"<h3>{site_name}爬取完成</h3><p>2026届相关新职位: {len(all_new_jobs)} 个</p><p>Excel生成失败</p>"
        )
    
    return list(existing_jobs.values())


# ============================
# 主函数
# ============================

def main():
    """程序入口函数"""
    logger.info(f"===== 开始招聘信息爬取任务 =====")
    logger.info(f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(f"目标范围: {START_PAGE}-{END_PAGE}页，筛选2026届相关职位")
    
    try:
        # 爬取校招信息
        campus_data = process_site(
            "校招",
            SITE_URL,
            DATA_FILE_CAMPUS,
            EXCEL_FILE_CAMPUS
        )
        
        # 爬取实习信息
        intern_data = process_site(
            "实习",
            SITE_URL_INTERNSHIP,
            DATA_FILE_INTERNSHIP,
            EXCEL_FILE_INTERNSHIP
        )
        
        # 输出统计结果
        logger.info(f"校招2026届相关职位总数: {len(campus_data)}")
        logger.info(f"实习2026届相关职位总数: {len(intern_data)}")
        logger.info("===== 所有任务完成 =====")
        
    except Exception as e:
        logger.error(f"主程序错误: {e}")
        # 发送错误通知
        send_email(
            subject="招聘爬取出错通知",
            body=f"<h2>爬取失败</h2><p>错误: {str(e)}</p><p>时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>"
        )


if __name__ == "__main__":
    main()
